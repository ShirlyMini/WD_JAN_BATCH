# open xl/spreadsheet/wookbook/ --> sheet-->cell-->row and col

import openpyxl
#################write data
# wb = openpyxl.Workbook()
# # wb.save("./test1.xlsx")
# # wb = openpyxl.load_workbook()# excel is already create
# sh = wb['Sheet']
#
# # sh.cell(1,1).value='data1'
# # sh.cell(1,2).value='dat2'
# # sh.cell(1,3).value='dat3'
# # sh.cell(1,4).value='dat4'
#
# for r in range(1,6):
#     for c in range(1,6):
#         sh.cell(r, c).value = 'data'
#
# wb.save("./test1.xlsx")

#####################read data

wb = openpyxl.load_workbook("./test1.xlsx")# excel is already create
sh = wb['Sheet']
print(sh.max_row)
print(sh.max_column)
print(sh.cell(1,1).value)
print(sh.cell(1,2).value)
print(sh.cell(1,3).value)
print(sh.cell(1,4).value)
